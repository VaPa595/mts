import csv
from openpyxl import load_workbook, Workbook
import pandas as pd

# import xlsxwriter
# You can't append to existing file with 'xlsxwriter'. What you can do, is read the file, write it to a new one, ' \
#        'and then append on top of that. You could use 'openpyxl'
# todo determine output
# todo try/catch


def init_csv_file(file, columns):
    with open(file, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columns)


def insert_row_csv_file(file, row):
    with open(file, mode='a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(row)


def init_xlsx(file, columns):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)  # Append Row Values (list)
    wb.save(file)


def insert_row_xlsx(file, row):
    wb = load_workbook(file)
    ws = wb.worksheets[0]  # Select First Worksheet
    # Append Row Values
    ws.append(row)  # list
    wb.save(file)


def json_file_to_csv(json_file, csv_file):
    df = pd.read_json(json_file)
    df.to_csv(csv_file)


def list_to_xlsx(file, data_list):
    wb = load_workbook(file)
    ws = wb.worksheets[0]  # Select First Worksheet
    for row in data_list:
        ws.append(row)  # Append Row Values (list)
    wb.save(file)


def row_to_xlsx(file, row):
    wb = load_workbook(file)
    ws = wb.worksheets[0]  # Select First Worksheet
    ws.append(row)  # Append Row Values (list)
    wb.save(file)
